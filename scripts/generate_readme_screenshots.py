# -*- coding: utf-8 -*-
"""README에 쓰는 스크린샷 2장을 생성한다.

이 PC엔 LibreOffice/Excel COM 자동화가 없어 실제 엑셀 캡처가 불가능하므로,
`sample_data/검증결과_예시.xlsx`를 openpyxl로 직접 열어 셀 값과 채우기색을 읽고
Pillow로 표를 다시 그려 PNG로 저장한다. 값/색상을 손으로 타이핑하지 않고 항상
워크북에서 실시간으로 읽으므로, 예시 데이터가 바뀌면 이 스크립트를 다시 실행하는
것만으로 스크린샷이 최신 상태로 갱신된다.

사용법: python scripts/generate_readme_screenshots.py
(요구: pip install -r requirements-dev.txt 로 Pillow 설치)
"""

import os

import openpyxl
from PIL import Image, ImageDraw, ImageFont

REPO_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SOURCE_XLSX = os.path.join(REPO_ROOT, "sample_data", "검증결과_예시.xlsx")
OUT_DIR = os.path.join(REPO_ROOT, "images")

FONT_PATH = r"C:\Windows\Fonts\malgun.ttf"
FONT_BOLD_PATH = r"C:\Windows\Fonts\malgunbd.ttf"

HEADER_BG = (0x30, 0x54, 0x96)
HEADER_FG = (0xFF, 0xFF, 0xFF)
RED_BG = (0xF8, 0xCB, 0xAD)
GREEN_BG = (0xE2, 0xEF, 0xDA)
WHITE = (0xFF, 0xFF, 0xFF)
BLACK = (0x20, 0x20, 0x20)
BORDER = (0xBF, 0xBF, 0xBF)
TITLE_COLOR = (0x30, 0x54, 0x96)

PAD_X = 10
PAD_Y = 8

DETAIL_DISPLAY_COLUMNS = [
    "전표번호(거래ID)", "계정과목", "차변금액", "대변금액",
    "통화", "외화금액", "거래처", "거래_최종판정",
]
DETAIL_COL_ALIGN = ["l", "l", "r", "r", "l", "r", "l", "l"]

RED_FILL_RGB = "00F8CBAD"
GREEN_FILL_RGB = "00E2EFDA"


def font(size, bold=False):
    return ImageFont.truetype(FONT_BOLD_PATH if bold else FONT_PATH, size)


def text_size(draw, s, f):
    l, t, r, b = draw.textbbox((0, 0), s, font=f)
    return r - l, b - t


def comma(n):
    if n is None or n == "":
        return ""
    return f"{n:,}" if isinstance(n, (int, float)) else str(n)


def render_table(title, headers, rows, row_fills, out_path, col_align=None,
                  base_font_size=15, footer_note=None):
    f_title = font(base_font_size + 5, bold=True)
    f_head = font(base_font_size, bold=True)
    f_body = font(base_font_size)
    f_note = font(base_font_size - 2)

    dummy = Image.new("RGB", (10, 10))
    d = ImageDraw.Draw(dummy)

    ncols = len(headers)
    col_align = col_align or ["l"] * ncols
    col_widths = []
    for c in range(ncols):
        w = text_size(d, str(headers[c]), f_head)[0]
        for row in rows:
            w = max(w, text_size(d, str(row[c]), f_body)[0])
        col_widths.append(w + PAD_X * 2)

    row_h = max(text_size(d, "가", f_body)[1], 18) + PAD_Y * 2
    head_h = max(text_size(d, "가", f_head)[1], 18) + PAD_Y * 2
    title_h = text_size(d, title, f_title)[1] + PAD_Y * 3 if title else 0
    note_h = (text_size(d, footer_note, f_note)[1] + PAD_Y * 2) if footer_note else 0

    total_w = sum(col_widths) + 2
    total_h = title_h + head_h + row_h * len(rows) + note_h + 2

    img = Image.new("RGB", (total_w, total_h), WHITE)
    draw = ImageDraw.Draw(img)

    y = 0
    if title:
        draw.text((PAD_X, PAD_Y), title, font=f_title, fill=TITLE_COLOR)
        y += title_h

    x = 0
    for c, htext in enumerate(headers):
        draw.rectangle([x, y, x + col_widths[c], y + head_h], fill=HEADER_BG)
        tw, th = text_size(draw, str(htext), f_head)
        tx = x + PAD_X if col_align[c] == "l" else x + col_widths[c] - tw - PAD_X
        draw.text((tx, y + (head_h - th) // 2 - 2), str(htext), font=f_head, fill=HEADER_FG)
        x += col_widths[c]
    y += head_h

    for r, row in enumerate(rows):
        fill = row_fills[r]
        x = 0
        for c, val in enumerate(row):
            draw.rectangle([x, y, x + col_widths[c], y + row_h], fill=fill, outline=BORDER)
            s = str(val)
            tw, th = text_size(draw, s, f_body)
            tx = x + PAD_X if col_align[c] == "l" else x + col_widths[c] - tw - PAD_X
            draw.text((tx, y + (row_h - th) // 2 - 1), s, font=f_body, fill=BLACK)
            x += col_widths[c]
        y += row_h

    if footer_note:
        gray = (0xF2, 0xF2, 0xF2)
        draw.rectangle([0, y, total_w, y + note_h], fill=gray, outline=BORDER)
        tw, th = text_size(draw, footer_note, f_note)
        draw.text(((total_w - tw) // 2, y + (note_h - th) // 2 - 1), footer_note,
                   font=f_note, fill=(0x60, 0x60, 0x60))

    img.save(out_path)
    print("saved", out_path, img.size)


def render_summary(wb):
    ws = wb["요약"]
    headers = ["항목", "값"]
    rows = []
    for r in range(3, ws.max_row + 1):
        label = ws.cell(row=r, column=1).value
        value = ws.cell(row=r, column=2).value
        if label is None:
            continue
        rows.append([label, comma(value)])
    fills = [WHITE] * len(rows)

    title = ws.cell(row=1, column=1).value
    render_table(
        title, headers, rows, fills,
        os.path.join(OUT_DIR, "screenshot_summary.png"),
        col_align=["l", "l"],
    )


def render_detail(wb):
    ws = wb["A.표본전체상세"]
    header_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    col_index = {name: header_row.index(name) + 1 for name in DETAIL_DISPLAY_COLUMNS}
    txn_col = col_index["전표번호(거래ID)"]

    red_total = 0
    green_total = 0
    seen_txn = set()
    picked_red, picked_green = [], []

    for r in range(2, ws.max_row + 1):
        fill_rgb = ws.cell(row=r, column=1).fill.fgColor.rgb
        is_red = fill_rgb == RED_FILL_RGB
        is_green = fill_rgb == GREEN_FILL_RGB
        if is_red:
            red_total += 1
        elif is_green:
            green_total += 1

        txn_id = ws.cell(row=r, column=txn_col).value
        if txn_id in seen_txn:
            continue
        seen_txn.add(txn_id)

        row_vals = [comma(ws.cell(row=r, column=col_index[name]).value)
                    for name in DETAIL_DISPLAY_COLUMNS]
        if is_red:
            picked_red.append(row_vals)
        elif is_green:
            picked_green.append(row_vals)

    total_data_rows = red_total + green_total
    shown_rows = picked_red + picked_green
    fills = [RED_BG] * len(picked_red) + [GREEN_BG] * len(picked_green)

    footer_note = (
        f"실제 {total_data_rows}행 중 이상치/부적정 {red_total}건 · 정상 {green_total}건 "
        f"(여기서는 거래별 대표 {len(shown_rows)}행만 표시)"
    )

    render_table(
        None, DETAIL_DISPLAY_COLUMNS, shown_rows, fills,
        os.path.join(OUT_DIR, "screenshot_detail.png"),
        col_align=DETAIL_COL_ALIGN,
        base_font_size=14,
        footer_note=footer_note,
    )
    return red_total, green_total, total_data_rows, len(shown_rows)


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    wb = openpyxl.load_workbook(SOURCE_XLSX, data_only=True)
    render_summary(wb)
    red_total, green_total, total, shown = render_detail(wb)
    print(f"A.표본전체상세: 전체 {total}행 (빨강 {red_total} / 초록 {green_total}), "
          f"대표 {shown}행 발췌")


if __name__ == "__main__":
    main()
